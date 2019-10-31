#!/usr/local/bin/python2.7
# encoding: utf-8
from openpyxl import Workbook
from openpyxl import Workbook
from openpyxl.compat import range
from openpyxl.cell import get_column_letter
from openpyxl import load_workbook
import time
import math
from openpyxl.styles import colors
from openpyxl.styles import Font, Color
from openpyxl.styles import colors
from util import checkFileValid, getCurrTime
from Tkconstants import INSERT


def loadData(fileName, logger, text_field):
    #print "start load data %s" % time.clock()
    logger.info("Start to load data from file: %s" % fileName)
    text_field.insert(INSERT, "[%s]: Start to load data from file: %s\n" %(getCurrTime(), fileName))
    
    quanWorkBook = {}
    QuanData = load_workbook(fileName)
    #print "end load data %s" % time.clock()
    logger.info("Finish loading data")
    text_field.insert(INSERT, "[%s]: Finish loading data\n" %(getCurrTime()))
    
    sheetNames = QuanData.get_sheet_names()
    currentSheetIndex = 1
    sheetNum = len(sheetNames)
    #print sheetNames
    sheetNames.sort()   
    for currentSheetName in sheetNames:
        #print "load sheet %s: %s" % (currentSheetName, time.clock())
        logger.info("Start to load data from sheet [%d/%d]: %s" % (currentSheetIndex, sheetNum, currentSheetName))
        
        text_field.insert(INSERT, "[%s]: Start to load data from sheet [%d/%d]: %s\n" %(getCurrTime(), currentSheetIndex, sheetNum, currentSheetName))
        currentSheetIndex += 1
        text_field.see("end")
        sheetData = QuanData.get_sheet_by_name(currentSheetName)
        rows = sheetData.rows;
        columns = sheetData.columns;
        rowsNum = len(rows)
        columsNum = len(columns)

        cnt = 0
        sheetRowTitleValue = []
        sheetColValue = []       
        #print "Data in ", currentSheetName, ":\n"
        for row in rows:
            for colValue in row:   
                             
                if(cnt < columsNum):
                    sheetRowTitleValue.append(colValue.value)
                else:
                    sheetColValue.append(colValue.value)
                    
                cnt = cnt + 1
        
        wholeColumData = []
        for i in range(columsNum):
            columnData = []
            for j in range(rowsNum - 1):
                index = j * columsNum + i;
                #print index
                columnData.append(sheetColValue[index])
            wholeColumData.append(columnData)
               
        
        #print wholeColumData
        
        sheetDataFinal = {}
        for i in range(columsNum):
            sheetDataFinal[sheetRowTitleValue[i]] = wholeColumData[i]
            
        #print sheetDataFinal
        
        quanWorkBook[currentSheetName] = sheetDataFinal
        
    logger.info("Finish loading all sheet's data")
    text_field.insert(INSERT, "[%s]: Finish loading all sheet's data\n" %(getCurrTime()))
    return quanWorkBook
            

def extractPartData(wordBook):
    
    wb = Workbook()
    #remove default worksheet
    wb.remove_sheet(wb.get_sheet_by_name("Sheet"))
    
    #wb.remove_sheet("Sheet")
    sheetNames = list(wordBook)
    
    for currentSheetName in sheetNames:
        ws = wb.create_sheet(title=currentSheetName)
        sheetData = wordBook[currentSheetName]
        sheetTitle = list(sheetData)
        
        ws.append(sheetTitle)
        
        #print "Data in sheet: ", currentSheetName, ":\n"
        allColumnData = []
        for currentTitle in sheetTitle:
            columnData = sheetData[currentTitle]
            allColumnData.append(columnData)
        
        rowsNum = len(allColumnData[0])
        columnsNum = len(sheetTitle)
        
        #print rowsNum, columnsNum
        
        #for debug file quan.xlsx
        rowsNum = 6
        for i in range(columnsNum):
            for j in range(rowsNum):
                ws.cell(row = j + 2, column = i + 1, value=allColumnData[i][j])
        
            #print columnData
            
            
        
        
    
    wb.save(filename="part.xlsx")
        


def extractColumnToFile(wordBook, outputFilename, logger):
    wb = Workbook()
    #remove default worksheet
    wb.remove_sheet(wb.get_sheet_by_name("Sheet"))
    
    expectedRT = "Expected RT"
    mzExpected = 'm/z (Expected)'
    compound = "Compound"
    area = "Area"
    actualRT = "Actual RT"
    formula = "Formula"
    
    
    #wb.remove_sheet("Sheet")
    sheetNames = list(wordBook)
    
    ws1 = wb.create_sheet("Sheet-1")
    ws2 = wb.create_sheet("Sheet-2")
    
    firstSheetHeader = [expectedRT, mzExpected, compound, formula]
    secondSheetHeader = [expectedRT, mzExpected, compound, formula]
    
    sheetNames.sort()
    
    for currentSheetName in sheetNames:
        firstSheetHeader.append(currentSheetName)
        secondSheetHeader.append(currentSheetName)
        
    ws1.append(firstSheetHeader)
    ws2.append(secondSheetHeader)
    
    sheetData = wordBook[sheetNames[0]];
    expectedRTData = sheetData[expectedRT]
    rowsNum = len(expectedRTData)
    
    mzExpectedData = sheetData[mzExpected]
    compoundData = sheetData[compound]
    
    formulaData = sheetData[formula]
    print("formulaData = " + str(formulaData))
    print("compoundData = " + str(compoundData))
    
    ft = Font(color=colors.RED)
    for i in range(rowsNum):
        ws1.cell(row = i + 2, column = 1, value = expectedRTData[i])
        ws1.cell(row = i + 2, column = 2, value = mzExpectedData[i])
        ws1.cell(row = i + 2, column = 3, value = compoundData[i])
        
        ws1.cell(row = i + 2, column = 4, value = formulaData[i])
        
        ws2.cell(row = i + 2, column = 1, value = expectedRTData[i])
        ws2.cell(row = i + 2, column = 2, value = mzExpectedData[i])
        ws2.cell(row = i + 2, column = 3, value = compoundData[i])
        ws2.cell(row = i + 2, column = 4, value = formulaData[i])
        #ws1.cell(row = i + 2, column = 3).font = ft
        
    columnIndex = 5
    
    for currentSheetName in sheetNames:
        sheetData = wordBook[currentSheetName]
        areaData = sheetData[area]
        actualRTData = sheetData[actualRT]
        
        
        for i in range(rowsNum):
            ws1.cell(row = i + 2, column = columnIndex, value = areaData[i])
            ws2.cell(row = i + 2, column = columnIndex, value = actualRTData[i])
            
            expectRTValue = expectedRTData[i]
            actualRTValue = actualRTData[i]
            
            actualRTOrigValue = actualRTValue
            if(actualRTValue == "N/F"):
                actualRTValue = float("0.0")
#             else:
#                 actualRTValue = float(str(actualRTValue))
                
            if(expectRTValue == "N/F"):
                expectRTValue = float("0.0")
#             else:
#                 expectRTValue = float(str(expectRTValue))
                
            diff = math.fabs(actualRTValue - expectRTValue)
            
            if(diff >= 0.2 and actualRTOrigValue != "N/F" ):
                ws2.cell(row = i + 2, column = columnIndex).font = ft
        
        
        columnIndex = columnIndex + 1
        
    
    wb.save(filename=outputFilename)
    logger.info("Finish extracting data, the result file is: %s" % outputFilename)
    #text_field.insert(INSERT, "[%s]: Finish loading all sheet's data\n" %(getCurrTime()))
    
    
    
    
    
# myExcelBook = loadData("quan.xlsx")
# from constant import *
# import sys
# from util import printDict
# rowTitleMust = [compoundNameTitle, mzExpectedTitle, libraryScoreTitle, measuredAreaTitle, ipTitle, lsTitle, rtMeasuredTitle, mzDeltaTitle]
# 
# [missingNum, missingRowTitleDict] = checkFileValid(myExcelBook, rowTitleMust)
# if(missingNum > 0):
#     print "Error: Some worksheet missing some column(s):"
#     printDict(missingRowTitleDict)
#     #sys.exit()
#     raise ValueError("Error: Some worksheet missing some column(s), see missing column above")
# #extractColumnToFile(myExcelBook)
# extractColumnToFile(myExcelBook, "quan-filter.xlsx")

#print myExcelBook






