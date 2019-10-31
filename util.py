
from openpyxl import Workbook
from openpyxl import Workbook
from openpyxl.compat import range
from openpyxl.cell import get_column_letter
from openpyxl import load_workbook
import time
import math
from openpyxl.styles import colors
from openpyxl.styles import Font, Color
from openpyxl.styles import colors
from constant import compoundNameTitle, libraryScoreTitle, rtMeasuredTitle,\
    NAValue, measuredAreaTitle, ipTitle, lsTitle, mzDeltaTitle


def loadData(fileName, logger):
    #print "start load data %s" % time.clock()
    logger.info("Start loading data: " + fileName)
    
    wholeWorkBook = {}
    inputDataDict = load_workbook(fileName)
    #print "end load data %s" % time.clock()
    logger.info("Finish loading data: " + fileName)
    
    sheetNames = inputDataDict.get_sheet_names()
    #print sheetNames
    sheetNames.sort()   
    for currentSheetName in sheetNames:
        #print "load sheet Name %s: %s" % (currentSheetName, time.clock())
        logger.info("Start to load sheet data: " + currentSheetName)
        sheetData = inputDataDict.get_sheet_by_name(currentSheetName)
        rows = sheetData.rows;
        columns = sheetData.columns;
        rowsNum = len(rows)
        columsNum = len(columns)

        cnt = 0
        sheetRowTitleValue = []
        sheetColValue = []       
        #print "Data in ", currentSheetName, ":\n"
        for row in rows:
            for colValue in row:   
                             
                if(cnt < columsNum):
                    sheetRowTitleValue.append(colValue.value)
                else:
                    sheetColValue.append(colValue.value)
                    
                cnt = cnt + 1
        
        wholeColumData = []
        for i in range(columsNum):
            columnData = []
            for j in range(rowsNum - 1):
                index = j * columsNum + i;
                #print index
                columnData.append(sheetColValue[index])
            wholeColumData.append(columnData)
               
        
        #print wholeColumData
        
        sheetDataFinal = {}
        for i in range(columsNum):
            sheetDataFinal[sheetRowTitleValue[i]] = wholeColumData[i]
            
        #print sheetDataFinal
        
        wholeWorkBook[currentSheetName] = sheetDataFinal
        
    return wholeWorkBook




def extractPartData(wordBook, rows = 6, output="partial.xlsx"):
    
    wb = Workbook()
    #remove default worksheet
    wb.remove_sheet(wb.get_sheet_by_name("Sheet"))
    
    #wb.remove_sheet("Sheet")
    sheetNames = list(wordBook)
    sheetNames.sort()
    
    
    for currentSheetName in sheetNames:
        ws = wb.create_sheet(title=currentSheetName)
        sheetData = wordBook[currentSheetName]
        sheetTitle = list(sheetData)
        
        print "sheet title = ", sheetTitle
        
        ws.append(sheetTitle)
        
        print "Data in sheet: ", currentSheetName, ":\n"
        allColumnData = []
        for currentTitle in sheetTitle:
            columnData = sheetData[currentTitle]
            allColumnData.append(columnData)
        
        rowsNum = len(allColumnData[0])
        columnsNum = len(sheetTitle)
        
        print rowsNum, columnsNum
        
        #for debug file quan.xlsx
        rowsNum = rows
        for i in range(columnsNum):
            for j in range(rowsNum):
                ws.cell(row = j + 2, column = i + 1, value=allColumnData[i][j])
        
            #print columnData
    
    
    wb.save(filename=output)
  
    

def createOutputWordBook():
    wb = Workbook()
    #remove default worksheet
    wb.remove_sheet(wb.get_sheet_by_name("Sheet"))
    
    return wb

def writeWordBook(wordbook, output="newbook.xlsx"):
    wordbook.save(filename = output)
    
def writeDataToColumn(wordBook, sheetName, data, columnIndex, rowStartIndex):
    ws = wordBook.get_sheet_by_name(sheetName)
    
    for i in range(len(data)):
        ws.cell(row = rowStartIndex + i, column = columnIndex, value = data[i])
        
def writeDataToRow(wordBook, sheetName, data, rowIndex, columnStartIndex):
    ws = wordBook.get_sheet_by_name(sheetName)
    
    for i in range(len(data)):
        ws.cell(row = rowIndex, column = columnStartIndex + i, value = data[i])

def highLightCell(wordBook, sheetName, rtValue, data, rowIndex, columnStartIndex):
    ws = wordBook.get_sheet_by_name(sheetName)
    ft = Font(color=colors.RED)
    
#     if(diff >= 0.2):
#                 ws2.cell(row = i + 2, column = columnIndex).font = ft
    for i in range(len(data)):
        if(data[i] == ""):
            continue
        diff = math.fabs(data[i] - rtValue)
        if(diff >= 0.2):
            ws.cell(row = rowIndex, column = columnStartIndex + i).font = ft
        

def extractMZExpectData(screenDataBook):
       
    mzExpectInfo = {}
    sheetNames = list(screenDataBook)
    compoundNameTitle = "Compound Name"
    mzExpectedTitle = "m/z (Expected)"
    for currSheetName in sheetNames:
        sheetData = screenDataBook[currSheetName]
        compoundNameData = sheetData[compoundNameTitle]
        mzExpectData = sheetData[mzExpectedTitle]
        
        for i in range(len(compoundNameData)):
            mzExpectInfo[compoundNameData[i]] = mzExpectData[i]
            
    
    return mzExpectInfo
 
def extractFormulaData(screenDataBook):
       
    formulaInfo = {}
    sheetNames = list(screenDataBook)
    compoundNameTitle = "Compound Name"
    formulaTitle = "Formula"
    for currSheetName in sheetNames:
        sheetData = screenDataBook[currSheetName]
        compoundNameData = sheetData[compoundNameTitle]
        formulaData = sheetData[formulaTitle]
        
        for i in range(len(compoundNameData)):
            formulaInfo[compoundNameData[i]] = formulaData[i]
            
    
    return formulaInfo


def extractLibMatchNameData(screenDataBook):
       
    libMatchNameInfo = {}
    sheetNames = list(screenDataBook)
    compoundNameTitle = "Compound Name"
    libMatchNameTitle = "Lib Match Name"
    for currSheetName in sheetNames:
        sheetData = screenDataBook[currSheetName]
        compoundNameData = sheetData[compoundNameTitle]
        libMatchNameData = sheetData[libMatchNameTitle]
        
        for i in range(len(compoundNameData)):
            libMatchNameInfo[compoundNameData[i]] = libMatchNameData[i]
            
    
    return libMatchNameInfo
 
        
 
def extractLSMARTData(screenDataBook):
    rtLibraryScoreInfo = {}
     
    sheetNames = list(screenDataBook)
              
    
    for currSheetName in sheetNames:
        sheetData = screenDataBook[currSheetName]
        compoundNameData = sheetData[compoundNameTitle]
        for i in range(len(compoundNameData)):
            rtLibraryScoreInfo[compoundNameData[i]] = []
    
    for currSheetName in sheetNames:
        sheetData = screenDataBook[currSheetName]
        compoundNameData = sheetData[compoundNameTitle]
        libraryScoreData = sheetData.get(libraryScoreTitle)
        rtMeasuredData = sheetData[rtMeasuredTitle]
        measuredAreaData = sheetData[measuredAreaTitle]
         
        for i in range(len(compoundNameData)):            
                mzDeltaValue = libraryScoreData[i]
                if(mzDeltaValue == "N/A"):
                    mzDeltaValue = NAValue
                rtLibraryScoreInfo[compoundNameData[i]].append([currSheetName, mzDeltaValue, measuredAreaData[i],rtMeasuredData[i]])
                
    
    return rtLibraryScoreInfo

def isAllNA(data):
    
    flag = True
    for i in range(len(data)):
        if(data[i] != NAValue):
            flag = False
            break
    
    return flag


def extractCPLSData(screenDataBook):
    cplsInfo = {}
     
    sheetNames = list(screenDataBook)
              
    
    for currSheetName in sheetNames:
        sheetData = screenDataBook[currSheetName]
        compoundNameData = sheetData[compoundNameTitle]
        for i in range(len(compoundNameData)):
            cplsInfo[compoundNameData[i]] = []
    
    for currSheetName in sheetNames:
        sheetData = screenDataBook[currSheetName]
        compoundNameData = sheetData[compoundNameTitle]
        libraryScoreData = sheetData.get(libraryScoreTitle)
       
         
        for i in range(len(compoundNameData)):            
                mzDeltaValue = libraryScoreData[i]
                if(mzDeltaValue == "N/A"):
                    mzDeltaValue = NAValue
                cplsInfo[compoundNameData[i]].append([currSheetName, mzDeltaValue])
                
    
    return cplsInfo

def extractIPData(screenDataBook):
    ipInfo = {}
     
    sheetNames = list(screenDataBook)
              
    
    for currSheetName in sheetNames:
        sheetData = screenDataBook[currSheetName]
        compoundNameData = sheetData[compoundNameTitle]
        for i in range(len(compoundNameData)):
            ipInfo[compoundNameData[i]] = []
    
    for currSheetName in sheetNames:
        sheetData = screenDataBook[currSheetName]
        compoundNameData = sheetData[compoundNameTitle]
        ipData = sheetData.get(ipTitle)
       
         
        for i in range(len(compoundNameData)):            
                ipValue = ipData[i]
                ipInfo[compoundNameData[i]].append([currSheetName, ipValue])
                
    
    return ipInfo

def isContainPass(data):
    flag = False
    for i in range(len(data)):
        if(data[i] == "Pass"):
            flag = True
            break
    
    return flag

def extractLSData(screenDataBook):
    lsInfo = {}
     
    sheetNames = list(screenDataBook)
              
    
    for currSheetName in sheetNames:
        sheetData = screenDataBook[currSheetName]
        compoundNameData = sheetData[compoundNameTitle]
        for i in range(len(compoundNameData)):
            lsInfo[compoundNameData[i]] = []
    
    for currSheetName in sheetNames:
        sheetData = screenDataBook[currSheetName]
        compoundNameData = sheetData[compoundNameTitle]
        lsData = sheetData.get(lsTitle)
       
         
        for i in range(len(compoundNameData)):            
                lsValue = lsData[i]
                lsInfo[compoundNameData[i]].append([currSheetName, lsValue])
                
    
    return lsInfo



def extractRTMeasuredData(screenDataBook):
    rtMeasuredInfo = {}
     
    sheetNames = list(screenDataBook)
              
    
    for currSheetName in sheetNames:
        sheetData = screenDataBook[currSheetName]
        compoundNameData = sheetData[compoundNameTitle]
        for i in range(len(compoundNameData)):
            rtMeasuredInfo[compoundNameData[i]] = []
    
    for currSheetName in sheetNames:
        sheetData = screenDataBook[currSheetName]
        compoundNameData = sheetData[compoundNameTitle]
        rtMeasuredData = sheetData.get(rtMeasuredTitle)
       
         
        for i in range(len(compoundNameData)):            
                mzMeasuredValue = rtMeasuredData[i]
                rtMeasuredInfo[compoundNameData[i]].append([currSheetName, mzMeasuredValue])
                
    
    return rtMeasuredInfo

def extractMZDeltaData(screenDataBook):
    mzDeltaInfo = {}
     
    sheetNames = list(screenDataBook)
              
    
    for currSheetName in sheetNames:
        sheetData = screenDataBook[currSheetName]
        compoundNameData = sheetData[compoundNameTitle]
        for i in range(len(compoundNameData)):
            mzDeltaInfo[compoundNameData[i]] = []
    
    for currSheetName in sheetNames:
        sheetData = screenDataBook[currSheetName]
        compoundNameData = sheetData[compoundNameTitle]
        mzDeltaData = sheetData.get(mzDeltaTitle)
       
         
        for i in range(len(compoundNameData)):            
                mzDeltaValue = mzDeltaData[i]
                mzDeltaInfo[compoundNameData[i]].append([currSheetName, mzDeltaValue])
                
    
    return mzDeltaInfo


def isSheetInList(inputSheetName, data):
    
    flag = False
    for [sheetName, maValue] in data:
        if(inputSheetName == sheetName):
            flag = True
            break
    
    return flag
        


def extractMeasuredAreaData(screenDataBook):
    measuredAreaInfo = {}
     
    sheetNames = list(screenDataBook)
              
    
    for currSheetName in sheetNames:
        sheetData = screenDataBook[currSheetName]
        compoundNameData = sheetData[compoundNameTitle]
        for i in range(len(compoundNameData)):
            measuredAreaInfo[compoundNameData[i]] = []
            
    
    
    
    
    for currSheetName in sheetNames:
        sheetData = screenDataBook[currSheetName]
        compoundNameData = sheetData[compoundNameTitle]
        measuredAreaData = sheetData.get(measuredAreaTitle)
       
         
        for i in range(len(compoundNameData)):            
                measuredAreaValue = measuredAreaData[i]
                measuredAreaInfo[compoundNameData[i]].append([currSheetName, measuredAreaValue])
                #measuredAreaInfo[compoundNameData[i]].append({currSheetName:measuredAreaValue})
                
    
    #padding compound name value that not in that sheet
    cpNames = list(measuredAreaInfo)
    numOfSheetNames = len(sheetNames)
    for currCPName in cpNames:
        for currSheetName in sheetNames:
            snmaList = measuredAreaInfo[currCPName]#.append({currSheetName:""})
            if(len(snmaList) == numOfSheetNames):
                continue
            
            flag = isSheetInList(currSheetName, snmaList)
            
            if(flag == False):
                measuredAreaInfo[currCPName].append([currSheetName, ""])
    
    
    return measuredAreaInfo

def extractALLRTMeasuredData(screenDataBook):
    rtmInfo = {}
     
    sheetNames = list(screenDataBook)
              
    
    for currSheetName in sheetNames:
        sheetData = screenDataBook[currSheetName]
        compoundNameData = sheetData[compoundNameTitle]
        for i in range(len(compoundNameData)):
            rtmInfo[compoundNameData[i]] = []
    
    
    for currSheetName in sheetNames:
        sheetData = screenDataBook[currSheetName]
        compoundNameData = sheetData[compoundNameTitle]
        rtmData = sheetData.get(rtMeasuredTitle)
       
         
        for i in range(len(compoundNameData)):            
                measuredAreaValue = rtmData[i]
                rtmInfo[compoundNameData[i]].append([currSheetName, measuredAreaValue])
                #rtmInfo[compoundNameData[i]].append({currSheetName:measuredAreaValue})
                
    
    #padding compound name value that not in that sheet
    cpNames = list(rtmInfo)
    numOfSheetNames = len(sheetNames)
    for currCPName in cpNames:
        for currSheetName in sheetNames:
            snrtmList = rtmInfo[currCPName]#.append({currSheetName:""})
            if(len(snrtmList) == numOfSheetNames):
                continue
            
            flag = isSheetInList(currSheetName, snrtmList)
            
            if(flag == False):
                rtmInfo[currCPName].append([currSheetName, ""])
    
    
    return rtmInfo


def printDict(data):
    keys = list(data)
    keys.sort()
    
    for key in keys:
        print key, ":"
        value = data[key]
        for item in value:
            print "====>", item
        
# wordDataBook = loadData("screen.xlsx")
# extractPartData(wordDataBook, 10, "small-screen2.xlsx")


def checkFileValid(dataBook, rowTitleList):
    sheetNames = list(dataBook)
    
    
    missingRowTitle = {}
    missNum = 0
    for sheetName in sheetNames:
        missingRowTitle[sheetName] = []
        sheetData = dataBook[sheetName]
        
        for title in rowTitleList:
            if(sheetData.has_key(title)):
                continue
            else:
                missNum = missNum + 1
                missingRowTitle[sheetName].append(title)
                
            
    
    return [missNum, missingRowTitle]     


def getCurrTime():    
    st = time.localtime()
    year = st.tm_year
    month = st.tm_mon
    day = st.tm_mday
    hour = st.tm_hour
    miniute = st.tm_min
    sec = st.tm_sec
    
    strTime = str(year) + "-" + str("%02d" % month) + "-" + str("%02d" % day) + "-" + str("%02d" % hour) + ":" + str("%02d" % miniute) + ":" + str("%02d" % sec)
    
    return strTime



            
    